#!/usr/bin/env python3
"""
Google 批量搜索工具 — 从 Excel 读取搜索词，批量搜索，关键词过滤，导出结果
v5: PySide6 现代 UI + undetected-chromedriver 真实浏览器搜索
"""

import sys
import time
import threading
from pathlib import Path
from datetime import datetime
import urllib.parse

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget,
    QVBoxLayout, QHBoxLayout, QGridLayout,
    QGroupBox, QLabel, QLineEdit, QPushButton,
    QSpinBox, QDoubleSpinBox, QComboBox,
    QTreeWidget, QTreeWidgetItem, QProgressBar,
    QPlainTextEdit, QFileDialog, QMessageBox,
    QHeaderView, QSplitter, QFrame,
)
from PySide6.QtCore import Qt, QThread, Signal, QTimer, QSize
from PySide6.QtGui import QFont, QIcon, QPalette, QColor, QBrush

import openpyxl

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ── 主题 ──────────────────────────────────────────────
DARK_QSS = """
QMainWindow, QWidget {
    background-color: #1e1e2e;
    color: #cdd6f4;
    font-family: "Microsoft YaHei", "Segoe UI", sans-serif;
    font-size: 13px;
}
QGroupBox {
    background-color: #262640;
    border: 1px solid #45475a;
    border-radius: 10px;
    margin-top: 18px;
    padding: 16px 12px 12px 12px;
    font-weight: bold;
    font-size: 13px;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 14px;
    padding: 0 8px;
    color: #89b4fa;
    background-color: #262640;
    border-radius: 4px;
}
QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox {
    background-color: #313244;
    color: #cdd6f4;
    border: 1px solid #45475a;
    border-radius: 6px;
    padding: 5px 8px;
    min-height: 22px;
}
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus {
    border-color: #89b4fa;
}
QComboBox::drop-down {
    border: none;
    padding-right: 6px;
}
QComboBox QAbstractItemView {
    background-color: #313244;
    color: #cdd6f4;
    selection-background-color: #45475a;
    border: 1px solid #45475a;
    border-radius: 4px;
}
QPushButton {
    border-radius: 6px;
    padding: 7px 18px;
    font-weight: bold;
    border: none;
    min-height: 24px;
}
QPushButton:hover {
    opacity: 200;
}
QPushButton:pressed {
    padding-top: 8px;
    padding-bottom: 6px;
}
QPushButton#startBtn {
    background-color: #89b4fa;
    color: #1e1e2e;
}
QPushButton#startBtn:hover {
    background-color: #74c7ec;
}
QPushButton#startBtn:disabled {
    background-color: #45475a;
    color: #6c7086;
}
QPushButton#stopBtn {
    background-color: #f38ba8;
    color: #1e1e2e;
}
QPushButton#stopBtn:hover {
    background-color: #eba0ac;
}
QPushButton#stopBtn:disabled {
    background-color: #45475a;
    color: #6c7086;
}
QPushButton#exportBtn {
    background-color: #a6e3a1;
    color: #1e1e2e;
}
QPushButton#exportBtn:hover {
    background-color: #94e2d5;
}
QPushButton#exportBtn:disabled {
    background-color: #45475a;
    color: #6c7086;
}
QPushButton#browseBtn {
    background-color: #45475a;
    color: #cdd6f4;
    border: 1px solid #585b70;
}
QPushButton#browseBtn:hover {
    background-color: #585b70;
}
QProgressBar {
    background-color: #313244;
    border: none;
    border-radius: 8px;
    text-align: center;
    color: #cdd6f4;
    font-size: 12px;
    min-height: 18px;
}
QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #89b4fa, stop:1 #cba6f7);
    border-radius: 8px;
}
QTreeWidget {
    background-color: #1e1e2e;
    color: #cdd6f4;
    border: 1px solid #45475a;
    border-radius: 6px;
    alternate-background-color: #262640;
    gridline-color: #313244;
}
QTreeWidget::item {
    padding: 4px 4px;
    border-bottom: 1px solid #313244;
}
QTreeWidget::item:selected {
    background-color: #45475a;
    color: #cdd6f4;
}
QHeaderView::section {
    background-color: #313244;
    color: #89b4fa;
    padding: 6px 8px;
    border: none;
    border-bottom: 2px solid #45475a;
    font-weight: bold;
}
QPlainTextEdit {
    background-color: #11111b;
    color: #a6adc8;
    border: 1px solid #45475a;
    border-radius: 6px;
    font-family: "Consolas", "Cascadia Code", monospace;
    font-size: 11px;
}
QSplitter::handle {
    background-color: #45475a;
    width: 2px;
}
QScrollBar:vertical {
    background: #1e1e2e;
    width: 10px;
    border-radius: 5px;
}
QScrollBar::handle:vertical {
    background: #45475a;
    border-radius: 5px;
    min-height: 30px;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0;
}
QScrollBar:horizontal {
    background: #1e1e2e;
    height: 10px;
    border-radius: 5px;
}
QScrollBar::handle:horizontal {
    background: #45475a;
    border-radius: 5px;
    min-width: 30px;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0;
}
"""

LIGHT_QSS = """
QMainWindow, QWidget {
    background-color: #eff1f5;
    color: #4c4f69;
    font-family: "Microsoft YaHei", "Segoe UI", sans-serif;
    font-size: 13px;
}
QGroupBox {
    background-color: #ffffff;
    border: 1px solid #ccd0da;
    border-radius: 10px;
    margin-top: 18px;
    padding: 16px 12px 12px 12px;
    font-weight: bold;
    font-size: 13px;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 14px;
    padding: 0 8px;
    color: #1e66f5;
    background-color: #ffffff;
    border-radius: 4px;
}
QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox {
    background-color: #ffffff;
    color: #4c4f69;
    border: 1px solid #ccd0da;
    border-radius: 6px;
    padding: 5px 8px;
    min-height: 22px;
}
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus {
    border-color: #1e66f5;
}
QComboBox QAbstractItemView {
    background-color: #ffffff;
    color: #4c4f69;
    selection-background-color: #e6e9ef;
    border: 1px solid #ccd0da;
}
QPushButton {
    border-radius: 6px;
    padding: 7px 18px;
    font-weight: bold;
    border: none;
    min-height: 24px;
}
QPushButton#startBtn {
    background-color: #1e66f5;
    color: #ffffff;
}
QPushButton#startBtn:hover {
    background-color: #04a5e5;
}
QPushButton#startBtn:disabled {
    background-color: #ccd0da;
    color: #9ca0b0;
}
QPushButton#stopBtn {
    background-color: #d20f39;
    color: #ffffff;
}
QPushButton#stopBtn:hover {
    background-color: #e64553;
}
QPushButton#stopBtn:disabled {
    background-color: #ccd0da;
    color: #9ca0b0;
}
QPushButton#exportBtn {
    background-color: #40a02b;
    color: #ffffff;
}
QPushButton#exportBtn:hover {
    background-color: #179299;
}
QPushButton#exportBtn:disabled {
    background-color: #ccd0da;
    color: #9ca0b0;
}
QPushButton#browseBtn {
    background-color: #e6e9ef;
    color: #4c4f69;
    border: 1px solid #ccd0da;
}
QPushButton#browseBtn:hover {
    background-color: #ccd0da;
}
QProgressBar {
    background-color: #e6e9ef;
    border: none;
    border-radius: 8px;
    text-align: center;
    color: #4c4f69;
    font-size: 12px;
    min-height: 18px;
}
QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #1e66f5, stop:1 #ea76cb);
    border-radius: 8px;
}
QTreeWidget {
    background-color: #ffffff;
    color: #4c4f69;
    border: 1px solid #ccd0da;
    border-radius: 6px;
    alternate-background-color: #f4f4f9;
    gridline-color: #e6e9ef;
}
QTreeWidget::item {
    padding: 4px 4px;
    border-bottom: 1px solid #e6e9ef;
}
QTreeWidget::item:selected {
    background-color: #1e66f5;
    color: #ffffff;
}
QHeaderView::section {
    background-color: #e6e9ef;
    color: #4c4f69;
    padding: 6px 8px;
    border: none;
    border-bottom: 2px solid #ccd0da;
    font-weight: bold;
}
QPlainTextEdit {
    background-color: #f4f4f9;
    color: #5c5f77;
    border: 1px solid #ccd0da;
    border-radius: 6px;
    font-family: "Consolas", "Cascadia Code", monospace;
    font-size: 11px;
}
QSplitter::handle {
    background-color: #ccd0da;
    width: 2px;
}
QScrollBar:vertical {
    background: #eff1f5;
    width: 10px;
    border-radius: 5px;
}
QScrollBar::handle:vertical {
    background: #ccd0da;
    border-radius: 5px;
    min-height: 30px;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0;
}
QScrollBar:horizontal {
    background: #eff1f5;
    height: 10px;
    border-radius: 5px;
}
QScrollBar::handle:horizontal {
    background: #ccd0da;
    border-radius: 5px;
    min-width: 30px;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0;
}
"""


# ── 搜索线程 ──────────────────────────────────────────
class SearchWorker(QThread):
    log_msg = Signal(str)
    progress_update = Signal(int, str)
    result_found = Signal(str, str, str, bool)
    search_done = Signal(int, int)
    enable_buttons = Signal()

    def __init__(self, rows_data, num_results, delay, excel_info):
        super().__init__()
        self.rows_data = rows_data
        self.num_results = num_results
        self.delay = delay
        self.excel_info = excel_info
        self.is_running = True
        self.all_results = []

    def run(self):
        driver = None
        try:
            driver = self._init_driver()
        except Exception as e:
            self.log_msg.emit(f"❌ 启动浏览器失败: {e}")
            self.enable_buttons.emit()
            return

        total = len(self.rows_data)
        for i, (query, filter_kw) in enumerate(self.rows_data):
            if not self.is_running:
                break

            self.progress_update.emit(int((i + 1) / total * 100), f"{i+1}/{total}")
            self.log_msg.emit(f"🔍 [{i+1}/{total}] 搜索: {query}")

            include_kw = (
                [k.strip().lower() for k in filter_kw.replace("，", ",").split(",") if k.strip()]
                if filter_kw else []
            )

            try:
                results = self._google_search(driver, query, self.num_results)
            except Exception as e:
                self.log_msg.emit(f"⚠️ 搜索失败: {e}")
                time.sleep(self.delay)
                continue

            kept = 0
            for j, url in enumerate(results):
                title = f"结果 {j+1}"
                is_kept = True
                if include_kw:
                    is_kept = any(kw in url.lower() for kw in include_kw)

                self.all_results.append({"query": query, "title": title, "url": url, "kept": is_kept})
                if is_kept:
                    kept += 1
                self.result_found.emit(query, title, url, is_kept)

            msg = f"    → {len(results)} 条结果"
            if include_kw:
                msg += f", 保留 {kept} 条"
            self.log_msg.emit(msg)
            time.sleep(self.delay)

        try:
            driver.quit()
        except Exception:
            pass

        kept_total = sum(1 for r in self.all_results if r["kept"])
        self.log_msg.emit(f"✅ 搜索完成 — 共 {len(self.all_results)} 条, 保留 {kept_total} 条")
        self.search_done.emit(len(self.all_results), kept_total)
        self.enable_buttons.emit()

    def stop(self):
        self.is_running = False

    def _init_driver(self):
        self.log_msg.emit("🚀 正在启动 Chrome 浏览器...")
        options = uc.ChromeOptions()
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--no-first-run")
        options.add_argument("--no-default-browser-check")
        options.add_argument("--disable-gpu")
        driver = uc.Chrome(options=options, version_main=None)
        driver.set_page_load_timeout(30)
        self.log_msg.emit("✅ Chrome 浏览器已启动")
        return driver

    def _google_search(self, driver, query: str, num: int = 10, lang: str = "zh"):
        urls = []
        params = {"q": query, "num": min(num, 30), "hl": lang}
        search_url = f"https://www.google.com/search?{urllib.parse.urlencode(params)}"

        try:
            driver.get(search_url)
        except Exception as e:
            raise Exception(f"页面加载失败: {e}")

        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#search"))
            )
        except Exception:
            pt = driver.page_source[:2000].lower()
            if "captcha" in pt or "unusual traffic" in pt:
                raise Exception("Google 触发 CAPTCHA，请手动在浏览器中完成验证后重试")
            if "consent.google" in driver.current_url or "before you continue" in pt:
                raise Exception("Google 弹出同意页，请手动点击 Accept all 后重试")
            time.sleep(2)

        # 策略 1
        try:
            for link in driver.find_elements(By.CSS_SELECTOR, 'a[jsname="UWckNb"]'):
                href = link.get_attribute("href")
                if href and href.startswith("http") and "google.com" not in href:
                    if href not in urls:
                        urls.append(href)
                    if len(urls) >= num:
                        return urls
        except Exception:
            pass

        # 策略 2
        if not urls:
            try:
                for h3 in driver.find_elements(By.TAG_NAME, "h3"):
                    try:
                        a = h3.find_element("xpath", "./ancestor::a")
                        href = a.get_attribute("href")
                        if href and href.startswith("http") and "google.com" not in href:
                            if href not in urls:
                                urls.append(href)
                    except Exception:
                        continue
                    if len(urls) >= num:
                        return urls
            except Exception:
                pass

        # 策略 3
        if not urls:
            try:
                for link in driver.find_elements(By.CSS_SELECTOR, 'a[href^="http"]'):
                    href = link.get_attribute("href")
                    if href and not any(d in href for d in [
                        "google.com", "googleadservices.com", "youtube.com",
                        "accounts.google", "policies.google", "support.google",
                    ]):
                        if href not in urls:
                            urls.append(href)
                        if len(urls) >= num:
                            return urls
            except Exception:
                pass

        return urls


# ── 主窗口 ────────────────────────────────────────────
class GoogleBatchSearcher(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Google Batch Searcher")
        self.resize(960, 780)
        self.setMinimumSize(780, 620)
        self._center()

        self.worker = None
        self.all_results = []
        self.current_theme = "dark"

        self._build_ui()
        self._apply_theme("dark")

    def _center(self):
        screen = QApplication.primaryScreen().availableGeometry()
        x = (screen.width() - self.width()) // 2
        y = (screen.height() - self.height()) // 2
        self.move(x, y)

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(14, 10, 14, 10)
        root.setSpacing(8)

        # ── 顶栏 ──
        topbar = QHBoxLayout()
        title_lbl = QLabel("🔍  Google Batch Searcher")
        title_lbl.setStyleSheet("font-size: 18px; font-weight: bold; color: #89b4fa; padding: 4px 0;")
        topbar.addWidget(title_lbl)
        topbar.addStretch()

        theme_lbl = QLabel("主题")
        theme_lbl.setStyleSheet("font-size: 12px; color: #6c7086;")
        topbar.addWidget(theme_lbl)
        self.theme_cb = QComboBox()
        self.theme_cb.addItems(["dark", "light"])
        self.theme_cb.setFixedWidth(80)
        self.theme_cb.currentTextChanged.connect(self._on_theme_change)
        topbar.addWidget(self.theme_cb)
        root.addLayout(topbar)

        # ── 配置区 ──
        config_gb = QGroupBox("📂 输入配置")
        config_grid = QGridLayout(config_gb)
        config_grid.setSpacing(8)

        # Excel 文件
        config_grid.addWidget(QLabel("Excel 文件"), 0, 0)
        self.excel_path = QLineEdit()
        self.excel_path.setPlaceholderText("选择 Excel 文件...")
        config_grid.addWidget(self.excel_path, 0, 1)
        browse_btn = QPushButton("浏览")
        browse_btn.setObjectName("browseBtn")
        browse_btn.clicked.connect(self._browse_excel)
        config_grid.addWidget(browse_btn, 0, 2)

        # 列配置
        config_grid.addWidget(QLabel("Sheet"), 1, 0)
        self.sheet_edit = QLineEdit("Sheet1")
        self.sheet_edit.setMaximumWidth(100)
        config_grid.addWidget(self.sheet_edit, 1, 1)

        config_grid.addWidget(QLabel("搜索词列"), 2, 0)
        self.qcol_edit = QLineEdit("A")
        self.qcol_edit.setMaximumWidth(50)
        self.qcol_edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
        config_grid.addWidget(self.qcol_edit, 2, 1)

        config_grid.addWidget(QLabel("过滤列 (可选)"), 3, 0)
        self.fcol_edit = QLineEdit("B")
        self.fcol_edit.setMaximumWidth(50)
        self.fcol_edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
        config_grid.addWidget(self.fcol_edit, 3, 1)

        # 参数
        params_gb = QGroupBox("⚙️ 搜索参数")
        params_layout = QHBoxLayout(params_gb)
        params_layout.addWidget(QLabel("结果数"))
        self.num_spin = QSpinBox()
        self.num_spin.setRange(1, 30)
        self.num_spin.setValue(10)
        params_layout.addWidget(self.num_spin)
        params_layout.addSpacing(16)
        params_layout.addWidget(QLabel("间隔(秒)"))
        self.delay_spin = QDoubleSpinBox()
        self.delay_spin.setRange(1.0, 15.0)
        self.delay_spin.setSingleStep(0.5)
        self.delay_spin.setValue(3.0)
        params_layout.addWidget(self.delay_spin)
        params_layout.addStretch()

        # 配置区排列
        config_row = QHBoxLayout()
        config_row.addWidget(config_gb, 3)
        config_row.addWidget(params_gb, 1)
        root.addLayout(config_row)

        # ── 按钮栏 ──
        btn_row = QHBoxLayout()
        self.start_btn = QPushButton("🔍 开始批量搜索")
        self.start_btn.setObjectName("startBtn")
        self.start_btn.setMinimumHeight(36)
        self.start_btn.clicked.connect(self.do_search)
        btn_row.addWidget(self.start_btn)

        self.stop_btn = QPushButton("⏹ 停止")
        self.stop_btn.setObjectName("stopBtn")
        self.stop_btn.setMinimumHeight(36)
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self._stop)
        btn_row.addWidget(self.stop_btn)

        self.export_btn = QPushButton("📥 导出 Excel")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setMinimumHeight(36)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.do_export)
        btn_row.addWidget(self.export_btn)

        btn_row.addStretch()

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedWidth(220)
        self.progress_bar.setValue(0)
        btn_row.addWidget(self.progress_bar)

        self.progress_label = QLabel("就绪")
        self.progress_label.setStyleSheet("font-size: 11px; color: #6c7086;")
        btn_row.addWidget(self.progress_label)
        root.addLayout(btn_row)

        # ── 结果 + 日志 分割区 ──
        splitter = QSplitter(Qt.Orientation.Vertical)

        # 结果表格
        result_gb = QGroupBox("📋 搜索结果")
        result_layout = QVBoxLayout(result_gb)
        result_layout.setContentsMargins(8, 12, 8, 8)

        self.result_tree = QTreeWidget()
        self.result_tree.setHeaderLabels(["搜索词", "标题", "链接", "保留"])
        self.result_tree.setAlternatingRowColors(True)
        self.result_tree.setRootIsDecorated(False)
        self.result_tree.setSelectionBehavior(QTreeWidget.SelectionBehavior.SelectRows)
        self.result_tree.itemDoubleClicked.connect(self._copy_url)

        header = self.result_tree.header()
        header.setStretchLastSection(False)
        header.resizeSection(0, 110)
        header.resizeSection(1, 190)
        header.resizeSection(2, 420)
        header.resizeSection(3, 50)

        result_layout.addWidget(self.result_tree)
        splitter.addWidget(result_gb)

        # 日志
        log_gb = QGroupBox("📜 运行日志")
        log_layout = QVBoxLayout(log_gb)
        log_layout.setContentsMargins(8, 12, 8, 8)

        self.log_text = QPlainTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumBlockCount(500)
        log_layout.addWidget(self.log_text)
        splitter.addWidget(log_gb)

        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 1)
        root.addWidget(splitter)

    # ── 主题 ──
    def _on_theme_change(self, theme):
        self.current_theme = theme
        self._apply_theme(theme)

    def _apply_theme(self, theme):
        if theme == "dark":
            self.setStyleSheet(DARK_QSS)
        else:
            self.setStyleSheet(LIGHT_QSS)

    # ── 日志 ──
    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.appendPlainText(f"[{ts}] {msg}")

    # ── 文件浏览 ──
    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "",
            "Excel 文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )
        if path:
            self.excel_path.setText(path)

    # ── 列号转换 ──
    def _col_to_index(self, col_letter: str) -> int:
        col_letter = col_letter.strip().upper()
        r = 0
        for c in col_letter:
            r = r * 26 + (ord(c) - ord('A') + 1)
        return r - 1

    # ── 停止 ──
    def _stop(self):
        if self.worker:
            self.worker.stop()
        self._log("⏹ 用户停止了搜索")

    # ── 复制链接 ──
    def _copy_url(self, item, column):
        url = item.text(2)
        QApplication.clipboard().setText(url)
        self._log(f"📋 已复制: {url}")

    # ── 搜索 ──
    def do_search(self):
        excel_path = self.excel_path.text().strip()
        if not excel_path or not Path(excel_path).exists():
            QMessageBox.critical(self, "错误", "请先选择一个有效的 Excel 文件")
            return

        sheet_name = self.sheet_edit.text().strip() or "Sheet1"
        query_col = self.qcol_edit.text().strip() or "A"
        filter_col = self.fcol_edit.text().strip() or ""
        num_results = self.num_spin.value()
        delay = self.delay_spin.value()

        try:
            q_idx = self._col_to_index(query_col)
            f_idx = self._col_to_index(filter_col) if filter_col else None
        except Exception:
            QMessageBox.critical(self, "错误", f"列号格式错误: {query_col} / {filter_col}")
            return

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            if sheet_name not in wb.sheetnames:
                QMessageBox.critical(self, "错误",
                                     f"Sheet '{sheet_name}' 不存在。\n可用: {', '.join(wb.sheetnames)}")
                return
            ws = wb[sheet_name]
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法读取 Excel: {e}")
            return

        rows_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= q_idx:
                continue
            q = str(row[q_idx]).strip() if row[q_idx] else ""
            if not q:
                continue
            fk = ""
            if f_idx is not None and len(row) > f_idx and row[f_idx]:
                fk = str(row[f_idx]).strip()
            rows_data.append((q, fk))
        wb.close()

        if not rows_data:
            QMessageBox.information(self, "提示", "Excel 中没有找到搜索词（从第2行开始读取）")
            return

        self._log(f"📖 读取到 {len(rows_data)} 条搜索词")
        self._log(f"⚙️ 每条 {num_results} 个结果 | 间隔 {delay}s | 浏览器模式")

        # 清空
        self.all_results.clear()
        self.result_tree.clear()
        self.progress_bar.setValue(0)

        # 禁用按钮
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.export_btn.setEnabled(False)

        self.worker = SearchWorker(rows_data, num_results, delay, {})
        self.worker.log_msg.connect(self._log)
        self.worker.progress_update.connect(self._on_progress)
        self.worker.result_found.connect(self._on_result)
        self.worker.search_done.connect(self._on_search_done)
        self.worker.enable_buttons.connect(self._on_enable_buttons)
        self.worker.start()

    def _on_progress(self, pct, text):
        self.progress_bar.setValue(pct)
        self.progress_label.setText(text)

    def _on_result(self, query, title, url, is_kept):
        status = "✅" if is_kept else "⏭️"
        item = QTreeWidgetItem([query, title, url, status])
        self.result_tree.addTopLevelItem(item)

    def _on_search_done(self, total, kept):
        self.progress_label.setText(f"完成 {total} 条")

    def _on_enable_buttons(self):
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.export_btn.setEnabled(len(self.all_results) > 0)

    # ── 导出 ──
    def do_export(self):
        # 从 TreeWidget 重建 all_results（因为 worker 线程可能已经结束）
        all_items = []
        for i in range(self.result_tree.topLevelItemCount()):
            item = self.result_tree.topLevelItem(i)
            all_items.append({
                "query": item.text(0),
                "title": item.text(1),
                "url": item.text(2),
                "kept": item.text(3) == "✅",
            })

        kept_items = [r for r in all_items if r["kept"]]
        if not kept_items:
            QMessageBox.information(self, "提示", "没有可导出的结果")
            return

        out_path, _ = QFileDialog.getSaveFileName(
            self, "保存结果",
            f"google_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel 文件 (*.xlsx)"
        )
        if not out_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "搜索结果"
        ws.append(["搜索词", "结果链接", "标题", "是否保留"])
        for r in kept_items:
            ws.append([r["query"], r["url"], r["title"], "✅" if r["kept"] else "⏭️"])

        ws2 = wb.create_sheet("全部结果")
        ws2.append(["搜索词", "结果链接", "标题", "是否保留"])
        for r in all_items:
            ws2.append([r["query"], r["url"], r["title"], "✅" if r["kept"] else "⏭️"])

        wb.save(out_path)
        self._log(f"📥 已导出: {out_path}")
        QMessageBox.information(self, "导出成功",
                                f"结果已保存到:\n{out_path}\n\n{len(kept_items)} 条保留 / {len(all_items)} 条全部")


def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Google Batch Searcher")
    window = GoogleBatchSearcher()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
